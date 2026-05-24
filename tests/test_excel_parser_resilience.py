from __future__ import annotations

from openpyxl import Workbook

from app.parsers import parse_excel


def test_excel_parser_skips_leading_blank_rows_before_header(tmp_path):
    path = tmp_path / "leading_blank.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Framework"
    sheet.append([])
    sheet.append(["", "", ""])
    sheet.append(["Topic", "Description"])
    sheet.append(["Communication", "Use clear purpose, audience, and channel."])
    workbook.save(path)

    records = parse_excel(path)

    assert len(records) == 1
    assert "clear purpose" in records[0]["text"]
    assert records[0]["metadata"]["sheet_name"] == "Framework"


def test_excel_parser_falls_back_to_dense_sheet_text_without_header(tmp_path):
    path = tmp_path / "dense_notes.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Notes"
    sheet["B2"] = "Communication Framework"
    sheet["B4"] = "Purpose"
    sheet["C4"] = "Audience"
    sheet["B5"] = "Choose the right channel for the audience."
    workbook.save(path)

    records = parse_excel(path)

    assert len(records) == 1
    assert "Choose the right channel" in records[0]["text"]


def test_excel_parser_dense_fallback_for_one_column_notes(tmp_path):
    path = tmp_path / "one_column_notes.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Notes"
    sheet["A1"] = "Communication Framework"
    sheet["A3"] = "Use sender, receiver, message, channel, and feedback."
    workbook.save(path)

    records = parse_excel(path)

    assert len(records) == 1
    assert records[0]["metadata"]["excel_parse_mode"] == "dense_sheet"
    assert "Communication Framework" in records[0]["text"]
    assert "sender, receiver" in records[0]["text"]
