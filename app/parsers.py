"""
File parsers for CSV/TSV/Excel/PDF/HTML/TXT/XML/JSONL/NDJSON/JSON.

Each parser returns list[dict]: {"text": str, "metadata": dict}
"""

from __future__ import annotations

import csv
import json
import logging
import re
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.config import settings
from app.lang import detect_language

logger = logging.getLogger(__name__)

_TEXT_ENCODINGS = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]

_KB_TEXT_COLS = ("title", "content")
_KB_META_COLS = ("category", "keywords", "source_url")
_FAQ_TEXT_COLS = ("question", "answer")
_FAQ_META_COLS = ("category", "tags")
_STRUCTURED_RECORD_CONTAINER_KEYS = ("data", "items", "records", "results", "rows")

_BLOB_RE = re.compile(r"[{}<>]|<html|<div|<span|<br", re.IGNORECASE)
_OCR_WORD_RE = re.compile(r"[0-9A-Za-zÀ-ỹĐđ]{2,}", re.UNICODE)
_OCR_VI_CHAR_RE = re.compile(r"[À-ỹĐđ]", re.UNICODE)
_OCR_NOISE_RE = re.compile(r"[\\{}<>@#$%^*_~=]{1,}", re.UNICODE)


def _normalize_value(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none", "null"} else text


def _read_text_with_fallbacks(file_path: Path, encodings: list[str] | None = None) -> tuple[str, str]:
    last_err: Exception | None = None
    for encoding in encodings or _TEXT_ENCODINGS:
        try:
            return file_path.read_text(encoding=encoding), encoding
        except Exception as err:
            last_err = err
    raise ValueError(f"Cannot decode text file {file_path.name}: {last_err}")


def _collect_columns(rows: list[dict[str, Any]]) -> list[str]:
    columns: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row.keys():
            key_str = str(key)
            if key_str in seen:
                continue
            seen.add(key_str)
            columns.append(key_str)
    return columns


def _is_kb_format(columns: list[str]) -> bool:
    lowered = [col.lower() for col in columns]
    return "title" in lowered and "content" in lowered


def _is_faq_format(columns: list[str]) -> bool:
    lowered = [col.lower() for col in columns]
    return "question" in lowered and "answer" in lowered


def _score_column(values: list[str]) -> float:
    if not values:
        return 0.0

    non_empty = [value for value in values if value]
    if not non_empty:
        return 0.0

    numeric_re = re.compile(r"^\d+([.,]\d+)?$")
    string_ratio = sum(1 for value in non_empty if not numeric_re.match(value)) / len(non_empty)
    avg_len = min(sum(len(value) for value in non_empty) / len(non_empty), 500) / 500.0
    unique_ratio = len(set(non_empty)) / len(non_empty)

    score = string_ratio * 0.4 + avg_len * 0.4 + unique_ratio * 0.2

    blob_count = sum(1 for value in non_empty if _BLOB_RE.search(value))
    score -= (blob_count / len(non_empty)) * 0.5

    num_count = sum(1 for value in non_empty if numeric_re.match(value))
    score -= (num_count / len(non_empty)) * 0.3

    return max(score, 0.0)


def _detect_content_column(columns: list[str], rows: list[dict[str, Any]]) -> str | None:
    best_col: str | None = None
    best_score = -1.0
    for col in columns:
        values = [_normalize_value(row.get(col, "")) for row in rows]
        score = _score_column(values)
        if score > best_score:
            best_score = score
            best_col = col
    return best_col


def _build_kb_text(row: dict[str, Any], columns: list[str]) -> tuple[str, dict[str, str]]:
    lower_map = {col.lower(): col for col in columns}

    title = _normalize_value(row.get(lower_map.get("title", ""), ""))
    content = _normalize_value(row.get(lower_map.get("content", ""), ""))
    keywords = _normalize_value(row.get(lower_map.get("keywords", ""), ""))

    parts: list[str] = []
    if title:
        parts.append(f"{title}:")
    if content:
        parts.append(content)
    if keywords:
        parts.append(f"Keywords: {keywords}")

    metadata: dict[str, str] = {}
    for col in _KB_META_COLS:
        source_col = lower_map.get(col)
        if not source_col:
            continue
        value = _normalize_value(row.get(source_col, ""))
        if value:
            metadata[col] = value

    for col in columns:
        if col.lower() in _KB_TEXT_COLS or col.lower() in _KB_META_COLS:
            continue
        value = _normalize_value(row.get(col, ""))
        if value:
            metadata[col] = value

    return " ".join(parts).strip(), metadata


def _build_faq_text(row: dict[str, Any], columns: list[str]) -> tuple[str, dict[str, str]]:
    lower_map = {col.lower(): col for col in columns}

    question = _normalize_value(row.get(lower_map.get("question", ""), ""))
    answer = _normalize_value(row.get(lower_map.get("answer", ""), ""))
    tags = _normalize_value(row.get(lower_map.get("tags", ""), ""))
    category = _normalize_value(row.get(lower_map.get("category", ""), ""))

    parts: list[str] = []
    if question:
        parts.append(f"Question: {question}")
    if answer:
        parts.append(f"Answer: {answer}")
    if tags:
        parts.append(f"Keywords: {tags}")

    metadata: dict[str, str] = {}
    if category:
        metadata["category"] = category
    if tags:
        metadata["keywords"] = tags

    for col in columns:
        if col.lower() in _FAQ_TEXT_COLS or col.lower() in _FAQ_META_COLS:
            continue
        value = _normalize_value(row.get(col, ""))
        if value:
            metadata[col] = value

    return " | ".join(parts).strip(), metadata


def _build_generic_text(
    row: dict[str, Any],
    columns: list[str],
    content_col: str | None,
) -> tuple[str, dict[str, str]]:
    metadata: dict[str, str] = {}
    main_text = ""

    for col in columns:
        value = _normalize_value(row.get(col, ""))
        if not value:
            continue
        if col == content_col:
            main_text = value
        else:
            metadata[col] = value

    if not main_text:
        parts = [f"{col}: {_normalize_value(row.get(col, ''))}" for col in columns if _normalize_value(row.get(col, ""))]
        main_text = " | ".join(parts)

    return main_text.strip(), metadata


def _rows_to_records(
    rows: list[dict[str, Any]],
    columns: list[str],
    base_meta: dict[str, Any],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []

    kb = _is_kb_format(columns)
    faq = _is_faq_format(columns)
    content_col = None if (kb or faq) else _detect_content_column(columns, rows)

    for row_index, row in enumerate(rows, start=2):
        if kb:
            text, extra = _build_kb_text(row, columns)
        elif faq:
            text, extra = _build_faq_text(row, columns)
        else:
            text, extra = _build_generic_text(row, columns, content_col)

        text = text.strip()
        if not text:
            continue

        records.append(
            {
                "text": text,
                "metadata": {
                    **base_meta,
                    "row_num": row_index,
                    "columns": columns,
                    "lang": detect_language(text),
                    **extra,
                },
            }
        )

    return records


def _sniff_csv_dialect(sample: str) -> csv.Dialect | None:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        return None


def _read_delimited_rows(file_path: Path) -> tuple[list[dict[str, Any]], list[str], str, str]:
    last_err: Exception | None = None

    for encoding in _TEXT_ENCODINGS:
        try:
            with file_path.open("r", encoding=encoding, newline="") as fobj:
                sample = fobj.read(4096)
                fobj.seek(0)
                dialect = _sniff_csv_dialect(sample)
                reader = csv.DictReader(fobj, dialect=dialect) if dialect else csv.DictReader(fobj)
                columns = list(reader.fieldnames or [])
                rows = [{(key or ""): _normalize_value(value) for key, value in row.items()} for row in reader]
                delimiter = getattr(dialect, "delimiter", ",")
                return rows, columns, encoding, delimiter
        except Exception as err:
            last_err = err

    raise ValueError(f"Cannot decode delimited file: {last_err}")


def parse_csv(file_path: Path) -> list[dict[str, Any]]:
    rows, columns, encoding, delimiter = _read_delimited_rows(file_path)
    records = _rows_to_records(rows, columns, {"encoding": encoding, "delimiter": delimiter})
    logger.info(
        "Parsed delimited file: %s records from %s (encoding=%s delimiter=%r)",
        len(records),
        file_path.name,
        encoding,
        delimiter,
    )
    return records


def _row_non_empty_count(row: list[str]) -> int:
    return sum(1 for cell in row if cell)


def _looks_like_excel_header(row: list[str], following_rows: list[list[str]]) -> bool:
    non_empty = _row_non_empty_count(row)
    if non_empty < 2:
        return False
    unique_non_empty = {cell.lower() for cell in row if cell}
    if len(unique_non_empty) < non_empty:
        return False
    if not following_rows:
        return True
    rows_with_data = sum(1 for item in following_rows[:10] if _row_non_empty_count(item) > 0)
    return rows_with_data > 0


def _sheet_dense_records(sheet_title: str, normalized_rows: list[list[str]]) -> list[dict[str, Any]]:
    lines: list[str] = []
    for index, row in enumerate(normalized_rows, start=1):
        values = [cell for cell in row if cell]
        if values:
            lines.append(f"Row {index}: " + " | ".join(values))
    text = "\n".join(lines).strip()
    if not text:
        return []
    return [
        {
            "text": text,
            "metadata": {
                "sheet_name": sheet_title,
                "row_num": 1,
                "columns": [],
                "lang": detect_language(text),
                "excel_parse_mode": "dense_sheet",
            },
        }
    ]


def _parse_excel_with_openpyxl(file_path: Path, *, data_only: bool = True) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    workbook = load_workbook(filename=file_path, read_only=True, data_only=data_only)

    for sheet in workbook.worksheets:
        normalized_rows = [
            [_normalize_value(cell) for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
        non_empty_rows = [row for row in normalized_rows if any(row)]
        if not non_empty_rows:
            continue

        header_index: int | None = None
        for index, row in enumerate(normalized_rows):
            if _looks_like_excel_header(row, normalized_rows[index + 1 :]):
                header_index = index
                break

        if header_index is None:
            records.extend(_sheet_dense_records(sheet.title, normalized_rows))
            continue

        raw_cols = normalized_rows[header_index]
        normalized_columns = [
            col if col else f"col_{index + 1}"
            for index, col in enumerate(raw_cols)
        ]
        raw_rows: list[dict[str, Any]] = []
        for values in normalized_rows[header_index + 1 :]:
            row_map = {
                normalized_columns[index]: _normalize_value(values[index] if index < len(values) else "")
                for index in range(len(normalized_columns))
            }
            if any(row_map.values()):
                raw_rows.append(row_map)

        sheet_records = _rows_to_records(raw_rows, normalized_columns, {"sheet_name": sheet.title})
        if sheet_records:
            records.extend(sheet_records)
        else:
            records.extend(_sheet_dense_records(sheet.title, normalized_rows))

    workbook.close()
    return records


def _parse_excel_with_pandas_fallback(file_path: Path) -> list[dict[str, Any]]:
    try:
        import pandas as pd  # type: ignore
    except Exception as err:
        raise ValueError(
            "Legacy .xls requires optional dependency pandas/xlrd. "
            "Please convert to .xlsx or install pandas + xlrd."
        ) from err

    records: list[dict[str, Any]] = []
    excel_file = pd.ExcelFile(file_path)
    for sheet_name in excel_file.sheet_names:
        df = pd.read_excel(excel_file, sheet_name=sheet_name, dtype=str, keep_default_na=False)
        columns = list(df.columns)
        raw_rows = [{col: _normalize_value(row[col]) for col in columns} for _, row in df.iterrows()]
        records.extend(_rows_to_records(raw_rows, columns, {"sheet_name": sheet_name}))

    return records


def parse_excel(file_path: Path) -> list[dict[str, Any]]:
    try:
        records = _parse_excel_with_openpyxl(file_path)
        if not records:
            records = _parse_excel_with_openpyxl(file_path, data_only=False)
        logger.info("Parsed Excel via openpyxl: %s records from %s", len(records), file_path.name)
        return records
    except InvalidFileException:
        records = _parse_excel_with_pandas_fallback(file_path)
        logger.info("Parsed Excel via pandas fallback: %s records from %s", len(records), file_path.name)
        return records


def _pdf_text_pages(file_path: Path) -> dict[int, str]:
    from pdfminer.high_level import extract_pages
    from pdfminer.layout import LTTextContainer

    pages: dict[int, str] = {}
    for page_index, layout in enumerate(extract_pages(str(file_path)), start=1):
        texts: list[str] = []
        for element in layout:
            if isinstance(element, LTTextContainer):
                text = element.get_text().strip()
                if text:
                    texts.append(text)

        pages[page_index] = "\n".join(texts).strip()

    return pages


def _pdf_records_from_pages(pages: dict[int, str], *, extraction: str) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for page_index in sorted(pages):
        page_text = pages[page_index].strip()
        if not page_text:
            continue

        records.append(
            {
                "text": page_text,
                "metadata": {
                    "page_num": page_index,
                    "lang": detect_language(page_text),
                    "pdf_extraction": extraction,
                },
            }
        )
    return records


def _pdf_table_rows_to_records(
    table: list[list[Any]],
    *,
    page_num: int,
    table_idx: int,
) -> list[dict[str, Any]]:
    if not table:
        return []

    headers = [_normalize_value(cell) for cell in table[0]]
    if not any(headers):
        max_columns = max((len(row) for row in table if row), default=0)
        headers = [f"Column {index + 1}" for index in range(max_columns)]

    records: list[dict[str, Any]] = []
    for row_idx, row in enumerate(table[1:], start=2):
        values = [_normalize_value(cell) for cell in row]
        parts = []
        for index, value in enumerate(values):
            if not value:
                continue
            header = (
                headers[index]
                if index < len(headers) and headers[index]
                else f"Column {index + 1}"
            )
            parts.append(f"{header}: {value}")

        text = " | ".join(parts).strip()
        if not text:
            continue

        records.append(
            {
                "text": text,
                "metadata": {
                    "page_num": page_num,
                    "table_idx": table_idx,
                    "row_num": row_idx,
                    "lang": detect_language(text),
                    "pdf_extraction": "pdfplumber_table",
                },
            }
        )

    return records


def _pdf_table_records(file_path: Path) -> list[dict[str, Any]]:
    try:
        import pdfplumber
    except Exception:
        logger.info("PDF table extraction skipped for %s: pdfplumber is not installed", file_path.name)
        return []

    records: list[dict[str, Any]] = []
    try:
        with pdfplumber.open(str(file_path)) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables() or []
                for table_idx, table in enumerate(tables, start=1):
                    records.extend(
                        _pdf_table_rows_to_records(table, page_num=page_num, table_idx=table_idx)
                    )
    except Exception:
        logger.warning(
            "PDF table extraction failed for %s; continuing with text extraction",
            file_path.name,
            exc_info=True,
        )
        return []

    return records


def _ocr_pdf_pages(file_path: Path, page_numbers: list[int]) -> dict[int, str]:
    try:
        from pdf2image import convert_from_path
        import pytesseract
    except Exception as err:
        raise ValueError(
            "PDF OCR requires optional dependencies pdf2image and pytesseract. "
            "Install them with `pip install pdf2image pytesseract`, then install Poppler and Tesseract OCR."
        ) from err

    if settings.pdf_ocr_tesseract_cmd.strip():
        pytesseract.pytesseract.tesseract_cmd = settings.pdf_ocr_tesseract_cmd.strip()

    poppler_path = settings.pdf_ocr_poppler_path.strip() or None
    dpi = max(72, int(settings.pdf_ocr_dpi))
    timeout = max(1, int(settings.pdf_ocr_timeout_seconds))
    language = settings.pdf_ocr_language.strip() or "eng"
    max_pages = max(1, int(settings.pdf_ocr_max_pages))

    selected_pages = page_numbers[:max_pages]
    if len(page_numbers) > len(selected_pages):
        logger.warning(
            "PDF OCR page limit reached for %s: processing %s/%s pages",
            file_path.name,
            len(selected_pages),
            len(page_numbers),
        )

    ocr_pages: dict[int, str] = {}
    for page_number in selected_pages:
        try:
            images = convert_from_path(
                str(file_path),
                dpi=dpi,
                first_page=page_number,
                last_page=page_number,
                fmt="png",
                poppler_path=poppler_path,
            )
        except Exception as err:
            raise ValueError(
                "PDF OCR could not render pages. Install Poppler and set "
                "RAG_PDF_OCR_POPPLER_PATH if Poppler is not on PATH."
            ) from err

        page_texts: list[str] = []
        for image in images:
            try:
                page_texts.append(
                    pytesseract.image_to_string(
                        image,
                        lang=language,
                        config=_tesseract_base_config(),
                        timeout=timeout,
                    )
                )
            except Exception as err:
                raise ValueError(
                    "PDF OCR failed while running Tesseract. Install Tesseract OCR, "
                    "ensure the configured language data exists, or set RAG_PDF_OCR_TESSERACT_CMD."
                ) from err
        ocr_pages[page_number] = "\n".join(page_texts).strip()

    return ocr_pages


def parse_pdf(file_path: Path) -> list[dict[str, Any]]:
    text_pages = _pdf_text_pages(file_path)
    text_records = _pdf_records_from_pages(text_pages, extraction="pdfminer")
    table_records = _pdf_table_records(file_path)
    records = text_records + table_records
    total_text_chars = sum(len(record["text"]) for record in text_records)

    if not settings.pdf_ocr_enabled or total_text_chars >= max(0, settings.pdf_ocr_min_text_chars):
        logger.info(
            "Parsed PDF: %s records from %s (table_records=%s)",
            len(records),
            file_path.name,
            len(table_records),
        )
        return records

    min_page_chars = max(0, settings.pdf_ocr_min_text_chars)
    candidate_pages = [
        page_number
        for page_number in sorted(text_pages)
        if len(text_pages.get(page_number, "").strip()) < min_page_chars
    ]
    if not candidate_pages and not records:
        candidate_pages = [1]

    try:
        ocr_pages = _ocr_pdf_pages(file_path, candidate_pages)
    except ValueError:
        if records:
            logger.warning("PDF OCR failed for %s; returning non-OCR PDF records", file_path.name, exc_info=True)
            return records
        raise

    merged_pages = dict(text_pages)
    for page_number, ocr_text in ocr_pages.items():
        if ocr_text.strip():
            merged_pages[page_number] = ocr_text.strip()

    text_records = _pdf_records_from_pages(merged_pages, extraction="pdfminer_ocr")
    records = text_records + table_records
    logger.info(
        "Parsed PDF: %s records from %s (ocr_pages=%s, table_records=%s)",
        len(records),
        file_path.name,
        len([text for text in ocr_pages.values() if text.strip()]),
        len(table_records),
    )
    return records


def parse_html(file_path: Path) -> list[dict[str, Any]]:
    from bs4 import BeautifulSoup

    content, encoding = _read_text_with_fallbacks(file_path)
    soup = BeautifulSoup(content, "html.parser")

    for tag in soup(["script", "style", "nav", "footer", "header"]):
        tag.decompose()

    title = soup.title.string.strip() if soup.title and soup.title.string else ""
    body = soup.body if soup.body else soup
    text = body.get_text(separator="\n", strip=True).strip()
    if not text:
        return []

    return [{"text": text, "metadata": {"title": title, "lang": detect_language(text), "encoding": encoding}}]


def parse_text(file_path: Path) -> list[dict[str, Any]]:
    content, encoding = _read_text_with_fallbacks(file_path)
    content = content.strip()
    if not content:
        return []
    return [{"text": content, "metadata": {"title": file_path.stem, "lang": detect_language(content), "encoding": encoding}}]


def _ocr_image_text(image: Any) -> str:
    try:
        import pytesseract
        from PIL import ImageOps
    except Exception as err:
        raise ValueError(
            "Image OCR requires optional dependencies pytesseract and Pillow. "
            "Install them with `pip install pytesseract Pillow`, then install Tesseract OCR."
        ) from err

    if settings.pdf_ocr_tesseract_cmd.strip():
        pytesseract.pytesseract.tesseract_cmd = settings.pdf_ocr_tesseract_cmd.strip()

    language = settings.pdf_ocr_language.strip() or "eng"
    timeout = max(1, int(settings.pdf_ocr_timeout_seconds))
    base_config = _tesseract_base_config()

    gray = ImageOps.grayscale(image)
    autocontrast = ImageOps.autocontrast(gray)
    variants = [
        image,
        gray,
        gray.resize((gray.width * 2, gray.height * 2)),
        autocontrast.resize((autocontrast.width * 2, autocontrast.height * 2)),
        autocontrast.point(lambda value: 255 if value > 170 else 0).resize(
            (autocontrast.width * 2, autocontrast.height * 2)
        ),
    ]
    configs = ("--psm 6", "", "--psm 4", "--psm 11")
    candidates: list[str] = []
    for variant in variants:
        for config in configs:
            try:
                full_config = f"{base_config} {config}".strip()
                candidate = pytesseract.image_to_string(
                    variant,
                    lang=language,
                    config=full_config,
                    timeout=timeout,
                )
            except Exception as err:
                if not candidates:
                    raise ValueError(
                        "Image OCR failed while running Tesseract. Install Tesseract OCR, "
                        "ensure the configured language data exists, or set RAG_PDF_OCR_TESSERACT_CMD."
                    ) from err
                logger.debug("Image OCR fallback failed for config=%s", config, exc_info=True)
                continue
            if candidate.strip():
                candidates.append(candidate)

    if candidates:
        return max(candidates, key=_ocr_candidate_score)

    return ""


def _tesseract_base_config() -> str:
    tessdata_dir = settings.tesseract_data_dir.strip()
    if not tessdata_dir:
        return ""
    return f"--tessdata-dir {Path(tessdata_dir).resolve()}"


def _ocr_candidate_score(text: str) -> float:
    stripped = text.strip()
    if not stripped:
        return 0.0

    words = _OCR_WORD_RE.findall(stripped)
    vi_chars = _OCR_VI_CHAR_RE.findall(stripped)
    noise = _OCR_NOISE_RE.findall(stripped)
    one_char_lines = sum(1 for line in stripped.splitlines() if len(line.strip()) == 1)
    alpha_chars = sum(1 for char in stripped if char.isalpha())
    printable_chars = sum(1 for char in stripped if not char.isspace())
    alpha_ratio = alpha_chars / max(1, printable_chars)

    return (
        len(words) * 8.0
        + len(stripped) * 0.35
        + len(vi_chars) * 1.5
        + alpha_ratio * 40.0
        - len(noise) * 8.0
        - one_char_lines * 12.0
    )


def parse_image(file_path: Path) -> list[dict[str, Any]]:
    try:
        from PIL import Image, UnidentifiedImageError
    except ImportError as err:
        raise ValueError("Image OCR requires Pillow. Install with: pip install Pillow") from err

    try:
        with Image.open(file_path) as image:
            image.load()
            image_format = image.format or file_path.suffix.lstrip(".").upper()
            width, height = image.size
            text = _ocr_image_text(image).strip()
            region_texts = _ocr_image_region_texts(image, context_text=text)
    except (UnidentifiedImageError, OSError) as err:
        raise ValueError(f"Cannot open image file {file_path.name}") from err

    if not text:
        raise ValueError(
            "Image OCR produced no text. The image may not contain readable text, "
            "may need a clearer crop, or may require installing the matching Tesseract language data."
        )

    records = [
        {
            "text": text,
            "metadata": {
                "title": file_path.stem,
                "lang": detect_language(text),
                "image_format": image_format,
                "image_width": width,
                "image_height": height,
                "ocr_extraction": "image_ocr",
            },
        }
    ]
    for region_idx, region_text in enumerate(region_texts, start=1):
        records.append(
            {
                "text": region_text,
                "metadata": {
                    "title": file_path.stem,
                    "lang": detect_language(region_text),
                    "image_format": image_format,
                    "image_width": width,
                    "image_height": height,
                    "ocr_region_idx": region_idx,
                    "ocr_extraction": "image_ocr_region",
                },
            }
        )
    return records


def _ocr_image_region_texts(image: Any, *, context_text: str) -> list[str]:
    try:
        import cv2
        import numpy as np
        from PIL import Image
    except Exception:
        return []

    try:
        rgb = image.convert("RGB")
        source = cv2.cvtColor(np.array(rgb), cv2.COLOR_RGB2BGR)
        height, width = source.shape[:2]
        ratio = height / 500.0
        resized = cv2.resize(source, (max(1, int(width / ratio)), 500))
        gray = cv2.cvtColor(resized, cv2.COLOR_BGR2GRAY)
        gray = cv2.GaussianBlur(gray, (5, 5), 0)
        edged = cv2.Canny(gray, 50, 150)
        contours, _ = cv2.findContours(edged, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    except Exception:
        logger.debug("Image OCR region detection failed", exc_info=True)
        return []

    context = _ocr_context_prefix(context_text)
    region_texts: list[str] = []
    seen_boxes: set[tuple[int, int, int, int]] = set()
    min_area = width * height * 0.015
    for contour in sorted(contours, key=cv2.contourArea, reverse=True)[:12]:
        perimeter = cv2.arcLength(contour, True)
        approx = cv2.approxPolyDP(contour, 0.02 * perimeter, True)
        if len(approx) != 4:
            continue

        quad = approx.reshape(4, 2).astype("float32") * ratio
        if cv2.contourArea(quad.astype("float32")) < min_area:
            continue

        x, y, w, h = cv2.boundingRect(quad.astype("int32"))
        box_key = (round(x / 20), round(y / 20), round(w / 20), round(h / 20))
        if box_key in seen_boxes:
            continue
        seen_boxes.add(box_key)

        warped = _warp_quad(source, quad)
        if warped is None:
            continue
        try:
            region_image = Image.fromarray(cv2.cvtColor(warped, cv2.COLOR_BGR2RGB))
            region_text = _ocr_image_text(region_image).strip()
        except Exception:
            logger.debug("Image OCR region text extraction failed", exc_info=True)
            continue
        if len(region_text) < 20:
            continue

        region_texts.append(f"{context}\n{region_text}".strip() if context else region_text)
        if len(region_texts) >= 3:
            break

    return region_texts


def _warp_quad(source: Any, quad: Any) -> Any | None:
    import cv2
    import numpy as np

    sums = quad.sum(axis=1)
    diffs = np.diff(quad, axis=1).reshape(-1)
    rect = np.array(
        [
            quad[np.argmin(sums)],
            quad[np.argmin(diffs)],
            quad[np.argmax(sums)],
            quad[np.argmax(diffs)],
        ],
        dtype="float32",
    )
    top_left, top_right, bottom_right, bottom_left = rect
    target_width = max(
        np.linalg.norm(bottom_right - bottom_left),
        np.linalg.norm(top_right - top_left),
    )
    target_height = max(
        np.linalg.norm(top_right - bottom_right),
        np.linalg.norm(top_left - bottom_left),
    )
    if target_width < 80 or target_height < 40:
        return None

    target = np.array(
        [
            [0, 0],
            [target_width - 1, 0],
            [target_width - 1, target_height - 1],
            [0, target_height - 1],
        ],
        dtype="float32",
    )
    matrix = cv2.getPerspectiveTransform(rect, target)
    return cv2.warpPerspective(source, matrix, (int(target_width), int(target_height)))


def _ocr_context_prefix(text: str) -> str:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    context_lines = []
    for line in lines:
        upper = line.upper()
        if "CHÍNH" in upper or "SÁCH" in upper or "POLICY" in upper:
            context_lines.append(line)
        elif len(line) <= 80 and sum(1 for char in line if char.isupper()) >= max(4, len(line) // 3):
            context_lines.append(line)
        if len(context_lines) >= 3:
            break
    return "\n".join(context_lines)


def _flatten_scalar(value: Any) -> str:
    if isinstance(value, str):
        return _normalize_value(value)
    if isinstance(value, (int, float, bool)) or value is None:
        return _normalize_value(value)
    return _normalize_value(json.dumps(value, ensure_ascii=False))


def _flatten_mapping(obj: Any, prefix: str = "") -> dict[str, Any]:
    if isinstance(obj, dict):
        flattened: dict[str, Any] = {}
        for key, value in obj.items():
            next_prefix = f"{prefix}.{key}" if prefix else str(key)
            flattened.update(_flatten_mapping(value, next_prefix))
        return flattened

    if isinstance(obj, list):
        flattened: dict[str, Any] = {}
        if not obj:
            flattened[prefix or "value"] = ""
            return flattened
        for index, value in enumerate(obj):
            next_prefix = f"{prefix}.{index}" if prefix else str(index)
            flattened.update(_flatten_mapping(value, next_prefix))
        return flattened

    return {prefix or "value": _flatten_scalar(obj)}


def _extract_structured_rows(payload: Any) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if isinstance(payload, list):
        return [_flatten_mapping(item) for item in payload], {}

    if not isinstance(payload, dict):
        return [_flatten_mapping(payload)], {}

    envelope = {
        key: _flatten_scalar(value)
        for key, value in payload.items()
        if not isinstance(value, (dict, list))
    }

    for key in _STRUCTURED_RECORD_CONTAINER_KEYS:
        value = payload.get(key)
        if isinstance(value, list):
            rows = []
            for item in value:
                flattened = _flatten_mapping(item)
                rows.append(
                    {
                        **{f"envelope.{item_key}": item_value for item_key, item_value in envelope.items()},
                        **flattened,
                    }
                )
            return rows, {"record_container": key}

    list_keys = [key for key, value in payload.items() if isinstance(value, list)]
    if len(list_keys) == 1:
        key = list_keys[0]
        rows = []
        for item in payload[key]:
            flattened = _flatten_mapping(item)
            rows.append(
                {
                    **{f"envelope.{item_key}": item_value for item_key, item_value in envelope.items()},
                    **flattened,
                }
            )
        return rows, {"record_container": key}

    return [_flatten_mapping(payload)], {}


def _xml_to_data(element: ET.Element) -> Any:
    children = list(element)
    text = (element.text or "").strip()

    if not children:
        if element.attrib:
            data = {f"@{key}": _normalize_value(value) for key, value in element.attrib.items()}
            if text:
                data["#text"] = _normalize_value(text)
            return data
        return _normalize_value(text)

    grouped: dict[str, list[Any]] = {}
    for child in children:
        grouped.setdefault(child.tag, []).append(_xml_to_data(child))

    data: dict[str, Any] = {f"@{key}": _normalize_value(value) for key, value in element.attrib.items()}
    for tag, values in grouped.items():
        data[tag] = values[0] if len(values) == 1 else values
    if text:
        data["#text"] = _normalize_value(text)
    return data


def parse_jsonl(file_path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    extra_meta: dict[str, Any] = {}
    detected_mode = "jsonl"
    last_err: Exception | None = None

    for encoding in _TEXT_ENCODINGS:
        try:
            content = file_path.read_text(encoding=encoding)
            stripped = content.strip()
            if stripped.startswith("[") or stripped.startswith("{"):
                payload = json.loads(stripped)
                rows, extra_meta = _extract_structured_rows(payload)
                detected_mode = "json"
            else:
                rows = [_flatten_mapping(json.loads(line)) for line in stripped.splitlines() if line.strip()]
            records = _rows_to_records(
                rows,
                _collect_columns(rows),
                {"structured_format": detected_mode, "encoding": encoding, **extra_meta},
            )
            logger.info("Parsed JSON/JSONL: %s records from %s", len(records), file_path.name)
            return records
        except Exception as err:
            last_err = err

    raise ValueError(f"Cannot decode JSON/JSONL file {file_path.name}: {last_err}")


def parse_xml(file_path: Path) -> list[dict[str, Any]]:
    content, encoding = _read_text_with_fallbacks(file_path)
    try:
        root = ET.fromstring(content)
    except ET.ParseError as err:
        raise ValueError(f"Cannot parse XML: {err}") from err

    payload = _xml_to_data(root)
    rows, extra_meta = _extract_structured_rows(payload)
    records = _rows_to_records(
        rows,
        _collect_columns(rows),
        {"structured_format": "xml", "xml_root": root.tag, "encoding": encoding, **extra_meta},
    )
    logger.info("Parsed XML: %s records from %s", len(records), file_path.name)
    return records


PARSERS = {
    "csv": parse_csv,
    "tsv": parse_csv,
    "excel": parse_excel,
    "pdf": parse_pdf,
    "html": parse_html,
    "text": parse_text,
    "jsonl": parse_jsonl,
    "ndjson": parse_jsonl,
    "json": parse_jsonl,
    "xml": parse_xml,
    "image": parse_image,
}


def parse_file(file_path: Path, parser_type: str) -> list[dict[str, Any]]:
    if parser_type == "docx":
        from app.parsers_docx import parse_docx

        return parse_docx(file_path)

    parser = PARSERS.get(parser_type)
    if not parser:
        raise ValueError(f"Unknown parser type: {parser_type}")
    return parser(file_path)
